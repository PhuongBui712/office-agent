"""Generate the sales.xlsx fixture used by the live smoke walkthrough.

Run from the BE repo root:
    uv run python tests/fixtures/_gen_sales_xlsx.py

Produces tests/fixtures/sales.xlsx with three sheets:
- Customers (PK: customer_id)
- Products  (PK: product_id)
- Sales    (FK -> customers.customer_id, FK -> products.product_id)
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook


def build_workbook() -> Workbook:
    wb = Workbook()
    # Default sheet -> Customers
    customers = wb.active
    customers.title = "Customers"
    customers.append(["customer_id", "name", "region", "tier"])
    regions = ["NA", "EU", "APAC"]
    tiers = ["bronze", "silver", "gold", "platinum"]
    for i in range(1, 31):
        customers.append([
            i,
            f"Customer #{i:02d}",
            regions[i % 3],
            tiers[i % 4],
        ])

    products = wb.create_sheet("Products")
    products.append(["product_id", "name", "category", "unit_price"])
    categories = ["Software", "Hardware", "Service", "Subscription"]
    for i in range(1, 11):
        products.append([
            1000 + i,
            f"Product {chr(64 + i)}",
            categories[i % 4],
            round(50 + i * 17.5, 2),
        ])

    sales = wb.create_sheet("Sales")
    sales.append([
        "sale_id", "date", "customer_id", "product_id",
        "quantity", "amount",
    ])
    start = date(2024, 1, 1)
    for i in range(1, 51):
        cust = ((i - 1) % 30) + 1
        prod = 1000 + ((i - 1) % 10) + 1
        qty = ((i % 5) + 1)
        # Mirror the Products price formula so amount stays self-consistent.
        unit_price = round(50 + ((prod - 1000)) * 17.5, 2)
        sales.append([
            i,
            start + timedelta(days=i),
            cust,
            prod,
            qty,
            round(qty * unit_price, 2),
        ])

    return wb


def main() -> None:
    out = Path(__file__).resolve().parent / "sales.xlsx"
    wb = build_workbook()
    wb.save(out)
    size = out.stat().st_size
    print(f"wrote {out} ({size} bytes)")


if __name__ == "__main__":
    main()
